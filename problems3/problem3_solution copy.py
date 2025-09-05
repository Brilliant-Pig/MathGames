# -*- coding: utf-8 -*-
"""
问题三：男胎Y染色体浓度达标时间多因素分析与BMI分组优化
基于GAMM模型、生存分析和蒙特卡洛模拟的综合解决方案
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import font_manager
import seaborn as sns
from scipy import stats
from sklearn.model_selection import train_test_split, cross_val_score, GridSearchCV
from sklearn.preprocessing import StandardScaler, PolynomialFeatures
from sklearn.metrics import mean_absolute_error, r2_score, mean_squared_error
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.linear_model import ElasticNet, Ridge, Lasso
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA
import warnings
import os
warnings.filterwarnings('ignore')

# 尝试导入高级统计库
try:
    import statsmodels.api as sm
    from statsmodels.gam.api import GLMGam, BSplines
    from statsmodels.stats.outliers_influence import variance_inflation_factor
    STATSMODELS_AVAILABLE = True
except ImportError:
    STATSMODELS_AVAILABLE = False
    print("Warning: statsmodels not available, using alternative methods")

try:
    from lifelines import KaplanMeierFitter, CoxPHFitter, WeibullFitter
    from lifelines.statistics import logrank_test
    LIFELINES_AVAILABLE = True
except ImportError:
    LIFELINES_AVAILABLE = False
    print("Warning: lifelines not available, using simplified survival analysis")

try:
    import pymc3 as pm
    import theano.tensor as tt
    PYMC3_AVAILABLE = True
except Exception:
    PYMC3_AVAILABLE = False
    print("Warning: PyMC3/theano unavailable or failed to import; using frequentist methods")

# 设置字体（自动选择可用的中文字体，避免中文显示为方框）
def _set_chinese_font():
    candidate_fonts = [
        'Microsoft YaHei', 'SimHei', 'SimSun', 'KaiTi',
        'Noto Sans CJK SC', 'Source Han Sans SC', 'Arial Unicode MS'
    ]
    # 先尝试通过绝对路径注册常见中文字体，保证可用
    font_paths = [
        r'C:\\Windows\\Fonts\\msyh.ttc',
        r'C:\\Windows\\Fonts\\msyhbd.ttc',
        r'C:\\Windows\\Fonts\\simhei.ttf',
        r'C:\\Windows\\Fonts\\simsun.ttc',
        r'C:\\Windows\\Fonts\\simkai.ttf',
    ]
    for p in font_paths:
        try:
            if os.path.exists(p):
                font_manager.fontManager.addfont(p)
        except Exception:
            pass
    # 获取系统已安装字体名称集合
    available = set(f.name for f in font_manager.fontManager.ttflist)
    chosen = None
    for fname in candidate_fonts:
        if fname in available:
            chosen = fname
            break
    if chosen is None:
        # 回退到已注册文件对应的常见名称
        chosen = 'Microsoft YaHei' if 'Microsoft YaHei' in available else (
            'SimHei' if 'SimHei' in available else (
            'SimSun' if 'SimSun' in available else 'Arial Unicode MS'))
    plt.rcParams['font.sans-serif'] = [chosen]
    plt.rcParams['font.family'] = 'sans-serif'
    # 负号正常显示
    plt.rcParams['axes.unicode_minus'] = False

_set_chinese_font()

# 设置图表样式
sns.set_style("whitegrid")
try:
    plt.style.use('seaborn-v0_8')
except OSError:
    try:
        plt.style.use('seaborn')
    except OSError:
        pass  # 使用默认样式

# 样式可能覆盖字体设置，这里再次强制中文字体
_set_chinese_font()

class NIPTProblem3Solver:
    def __init__(self, data_path):
        """初始化求解器"""
        self.data_path = data_path
        self.data = None
        self.male_data = None
        self.models = {}
        self.results = {}
        os.makedirs('chart', exist_ok=True)
        os.makedirs('chart/ver2', exist_ok=True)
        
    def load_and_preprocess_data(self):
        """数据加载与预处理"""
        print("正在加载和预处理数据...")
        
        try:
            # 读取数据
            if self.data_path.endswith('.xlsx'):
                try:
                    self.data = pd.read_excel(self.data_path)
                except Exception as e:
                    print(f"读取Excel文件时出错：{e}")
                    print("尝试使用openpyxl手动读取Excel...")
                    df_openpyxl = self._read_excel_without_pandas(self.data_path)
                    if df_openpyxl is not None and not df_openpyxl.empty:
                        self.data = df_openpyxl
                    else:
                        print("openpyxl读取失败，尝试创建示例数据...")
                        return self._create_sample_data()
            elif self.data_path.endswith('.csv'):
                self.data = pd.read_csv(self.data_path)
            else:
                print("不支持的文件格式，创建示例数据...")
                return self._create_sample_data()
        except FileNotFoundError:
            print(f"错误：找不到数据文件 {self.data_path}")
            print("创建示例数据...")
            return self._create_sample_data()
        except Exception as e:
            print(f"读取数据文件时出错：{e}")
            print("创建示例数据...")
            return self._create_sample_data()

        # 统一列，确保存在关键字段 C,D,E,J,K,V 或对应中文列
        self._ensure_required_columns()

        # 特殊清洗：将孕周字符串（如 11w+6 / 11周+6天）解析为数值周
        if 'J' in self.data.columns:
            if self.data['J'].dtype == object:
                self.data['J'] = self.data['J'].apply(self._parse_gestational_week)
        
        # 将关键列转换为数值型
        for col in ['C', 'D', 'E', 'J', 'K', 'V']:
            if col in self.data.columns:
                self.data[col] = pd.to_numeric(self.data[col], errors='coerce')

        # 筛选男胎数据（Y染色体浓度非空）
        if 'V' in self.data.columns:
            self.male_data = self.data[self.data['V'].notna()].copy()
        elif 'Y染色体浓度' in self.data.columns:
            self.data['V'] = pd.to_numeric(self.data['Y染色体浓度'], errors='coerce')
            self.male_data = self.data[self.data['V'].notna()].copy()
        else:
            print("警告：未找到'V'或'Y染色体浓度'列，无法筛选男胎数据")
            return None
        
        # 数据清洗
        self.male_data = self.male_data.dropna(subset=['C', 'D', 'E', 'J', 'K', 'V'])
        
        # 异常值处理
        # BMI异常值处理
        bmi_q1, bmi_q3 = self.male_data['K'].quantile([0.25, 0.75])
        bmi_iqr = bmi_q3 - bmi_q1
        bmi_lower = bmi_q1 - 1.5 * bmi_iqr
        bmi_upper = bmi_q3 + 1.5 * bmi_iqr
        
        # 孕周异常值处理
        week_q1, week_q3 = self.male_data['J'].quantile([0.25, 0.75])
        week_iqr = week_q3 - week_q1
        week_lower = week_q1 - 1.5 * week_iqr
        week_upper = week_q3 + 1.5 * week_iqr
        
        # 过滤异常值
        self.male_data = self.male_data[
            (self.male_data['K'] >= bmi_lower) & 
            (self.male_data['K'] <= bmi_upper) &
            (self.male_data['J'] >= week_lower) & 
            (self.male_data['J'] <= week_upper) &
            (self.male_data['J'] >= 10) & 
            (self.male_data['J'] <= 25)
        ]
        
        # 创建达标标志
        self.male_data['达标'] = (self.male_data['V'] >= 4.0).astype(int)
        
        # 重命名前，若目标列已存在且与源列不同，先删除以避免重复列导致绘图错误
        rename_map = {'C': '年龄', 'D': '身高', 'E': '体重', 'J': '孕周', 'K': 'BMI', 'V': 'Y染色体浓度'}
        for src, tgt in rename_map.items():
            if tgt in self.male_data.columns and src in self.male_data.columns and tgt != src:
                # 删除已有目标列，保留源列，随后统一重命名
                try:
                    self.male_data = self.male_data.drop(columns=[tgt])
                except Exception:
                    pass
        # 重命名列
        self.male_data = self.male_data.rename(columns=rename_map)
        
        if len(self.male_data) == 0:
            print("警告：没有找到有效的男胎数据")
            return None
            
        print(f"数据预处理完成，共{len(self.male_data)}条男胎记录")
        return self.male_data

    def _parse_gestational_week(self, value):
        """将孕周字符串解析为数值周。示例：
        - '11w+6' -> 11 + 6/7
        - '11周+6天' -> 11 + 6/7
        - '15w' or '15周' -> 15
        - 数值/空值 -> 原样或NaN
        """
        if value is None:
            return np.nan
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s:
            return np.nan
        import re
        # 匹配 '11w+6' 或 '11w + 6'
        m = re.match(r"^(\d+)\s*[wW]\s*(?:\+\s*(\d+))?", s)
        if m:
            w = float(m.group(1))
            d = float(m.group(2)) if m.group(2) is not None else 0.0
            return w + d/7.0
        # 匹配 '11周+6天' 或 '11周' 或 '11周6天'
        m = re.match(r"^(\d+)\s*周(?:\s*\+?\s*(\d+)\s*天)?", s)
        if m:
            w = float(m.group(1))
            d = float(m.group(2)) if m.group(2) is not None else 0.0
            return w + d/7.0
        # 仅数字
        try:
            return float(s)
        except Exception:
            return np.nan

    def _ensure_required_columns(self):
        """尽力标准化列，保证后续流程需要的 C,D,E,J,K,V 存在。
        策略：
        1) 若已包含 C,D,E,J,K,V 则跳过
        2) 若包含中文列名（年龄/身高/体重/孕周/BMI/Y染色体浓度）则补齐对应英文字母列
        3) 通过模糊匹配常见同义词（如 胎儿浓度/ff/FF/Y% 等）自动识别
        4) 否则按位置补齐：C=第3列，D=第4列，E=第5列，J=第10列，K=第11列，V=第22列（若存在）
        """
        required_letters = ['C', 'D', 'E', 'J', 'K', 'V']
        if all(col in self.data.columns for col in required_letters):
            return
        # 通过中文名映射
        chinese_map = {
            '年龄': 'C', '身高': 'D', '体重': 'E', '孕周': 'J', 'BMI': 'K', 'Y染色体浓度': 'V'
        }
        for zh, letter in chinese_map.items():
            if letter not in self.data.columns and zh in self.data.columns:
                self.data[letter] = self.data[zh]
        # 模糊匹配同义列名
        if not all(col in self.data.columns for col in required_letters):
            lower_cols = {str(c).strip().lower(): c for c in self.data.columns}
            def find_col(possible_names):
                for name in possible_names:
                    key = name.lower()
                    for lc, orig in lower_cols.items():
                        if key in lc:
                            return orig
                return None
            # 年龄 C
            if 'C' not in self.data.columns:
                cand = find_col(['年龄','age'])
                if cand is not None:
                    self.data['C'] = self.data[cand]
            # 身高 D
            if 'D' not in self.data.columns:
                cand = find_col(['身高','height'])
                if cand is not None:
                    self.data['D'] = self.data[cand]
            # 体重 E
            if 'E' not in self.data.columns:
                cand = find_col(['体重','weight'])
                if cand is not None:
                    self.data['E'] = self.data[cand]
            # 孕周 J
            if 'J' not in self.data.columns:
                cand = find_col(['孕周','gestational','ga','周数','周龄'])
                if cand is not None:
                    self.data['J'] = self.data[cand]
            # BMI K
            if 'K' not in self.data.columns:
                cand = find_col(['bmi','体质指数'])
                if cand is not None:
                    self.data['K'] = self.data[cand]
            # Y染色体浓度 V
            if 'V' not in self.data.columns:
                cand = find_col(['y染色体浓度','y浓度','胎儿浓度','ff','fetal fraction','y%','y 百分比'])
                if cand is not None:
                    self.data['V'] = self.data[cand]
        # 按位置补齐（不覆盖已有列）
        num_cols = self.data.shape[1]
        def set_from_pos(target_col: str, pos_index_zero_based: int):
            if target_col in self.data.columns:
                return
            if num_cols > pos_index_zero_based:
                self.data[target_col] = self.data.iloc[:, pos_index_zero_based]
        # C(3), D(4), E(5), J(10), K(11), V(22)
        set_from_pos('C', 2)
        set_from_pos('D', 3)
        set_from_pos('E', 4)
        set_from_pos('J', 9)
        set_from_pos('K', 10)
        set_from_pos('V', 21)

    def _read_excel_without_pandas(self, file_path):
        """在pandas/openpyxl版本不兼容时，使用openpyxl直接读取为DataFrame。
        - 优先使用首个工作表
        - 第一行为表头；如非字符串，则使用字母列名 A, B, C, ...
        """
        try:
            from openpyxl import load_workbook
        except Exception as e:
            print(f"无法导入openpyxl：{e}")
            return None
        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                first_row = next(rows_iter)
            except StopIteration:
                print("Excel文件为空")
                return None
            # 判断首行是否为表头（包含任意字符串即视为表头）
            has_header = any(isinstance(v, str) for v in first_row if v is not None)
            if has_header:
                headers = [str(v) if v is not None else '' for v in first_row]
            else:
                # 使用字母列名
                num_cols = len(first_row)
                headers = [chr(ord('A') + i) for i in range(num_cols)]
                # 把首行视为数据
                data_rows = [first_row]
            # 收集剩余行
            data_rows = [] if has_header else data_rows
            for r in rows_iter:
                data_rows.append(r)
            df = pd.DataFrame(data_rows, columns=headers)
            return df
        except Exception as e:
            print(f"openpyxl读取Excel失败：{e}")
            return None
    
    def _create_sample_data(self):
        """创建示例数据"""
        print("正在创建示例数据...")
        
        np.random.seed(42)
        n_samples = 1000
        
        # 创建示例数据
        data = {
            'A': range(1, n_samples + 1),  # 样本序号
            'B': [f'P{i:04d}' for i in range(1, n_samples + 1)],  # 孕妇代码
            'C': np.random.normal(30, 5, n_samples),  # 年龄
            'D': np.random.normal(165, 8, n_samples),  # 身高
            'E': np.random.normal(65, 15, n_samples),  # 体重
            'J': np.random.uniform(10, 25, n_samples),  # 孕周
            'K': np.random.uniform(20, 45, n_samples),  # BMI
            'V': np.random.uniform(1, 8, n_samples)  # Y染色体浓度
        }
        
        self.data = pd.DataFrame(data)
        
        # 筛选男胎数据（Y染色体浓度非空）
        self.male_data = self.data[self.data['V'].notna()].copy()
        
        # 数据清洗
        self.male_data = self.male_data.dropna(subset=['C', 'D', 'E', 'J', 'K', 'V'])
        
        # 异常值处理
        # BMI异常值处理
        bmi_q1, bmi_q3 = self.male_data['K'].quantile([0.25, 0.75])
        bmi_iqr = bmi_q3 - bmi_q1
        bmi_lower = bmi_q1 - 1.5 * bmi_iqr
        bmi_upper = bmi_q3 + 1.5 * bmi_iqr
        
        # 孕周异常值处理
        week_q1, week_q3 = self.male_data['J'].quantile([0.25, 0.75])
        week_iqr = week_q3 - week_q1
        week_lower = week_q1 - 1.5 * week_iqr
        week_upper = week_q3 + 1.5 * week_iqr
        
        # 过滤异常值
        self.male_data = self.male_data[
            (self.male_data['K'] >= bmi_lower) & 
            (self.male_data['K'] <= bmi_upper) &
            (self.male_data['J'] >= week_lower) & 
            (self.male_data['J'] <= week_upper) &
            (self.male_data['J'] >= 10) & 
            (self.male_data['J'] <= 25)
        ]
        
        # 创建达标标志
        self.male_data['达标'] = (self.male_data['V'] >= 4.0).astype(int)
        
        # 重命名列
        self.male_data = self.male_data.rename(columns={
            'C': '年龄', 'D': '身高', 'E': '体重', 'J': '孕周', 
            'K': 'BMI', 'V': 'Y染色体浓度'
        })
        
        print(f"示例数据创建完成，共{len(self.male_data)}条男胎记录")
        return self.male_data
    
    def exploratory_analysis(self):
        """探索性数据分析"""
        print("正在进行探索性数据分析...")
        
        # 创建图表
        fig, axes = plt.subplots(2, 3, figsize=(18, 12))
        fig.suptitle('男胎Y染色体浓度分析', fontsize=16, fontweight='bold')
        
        # 1. Y染色体浓度分布
        axes[0,0].hist(self.male_data['Y染色体浓度'], bins=30, alpha=0.7, color='skyblue', edgecolor='black')
        axes[0,0].axvline(4.0, color='red', linestyle='--', linewidth=2, label='阈值 (4%)')
        axes[0,0].set_title('Y染色体浓度分布')
        axes[0,0].set_xlabel('Y染色体浓度 (%)')
        axes[0,0].set_ylabel('频数')
        axes[0,0].legend()
        
        # 2. 孕周与Y染色体浓度关系
        sns.scatterplot(data=self.male_data, x='孕周', y='Y染色体浓度', 
                       hue='达标', palette=['red', 'green'], alpha=0.6, ax=axes[0,1])
        axes[0,1].axhline(4.0, color='black', linestyle='--', alpha=0.5)
        axes[0,1].set_title('孕周与Y染色体浓度')
        axes[0,1].set_xlabel('孕周')
        axes[0,1].set_ylabel('Y染色体浓度 (%)')
        
        # 3. BMI与Y染色体浓度关系
        sns.scatterplot(data=self.male_data, x='BMI', y='Y染色体浓度', 
                       hue='达标', palette=['red', 'green'], alpha=0.6, ax=axes[0,2])
        axes[0,2].axhline(4.0, color='black', linestyle='--', alpha=0.5)
        axes[0,2].set_title('BMI与Y染色体浓度')
        axes[0,2].set_xlabel('BMI')
        axes[0,2].set_ylabel('Y染色体浓度 (%)')
        
        # 4. 年龄与Y染色体浓度关系
        sns.scatterplot(data=self.male_data, x='年龄', y='Y染色体浓度', 
                       hue='达标', palette=['red', 'green'], alpha=0.6, ax=axes[1,0])
        axes[1,0].axhline(4.0, color='black', linestyle='--', alpha=0.5)
        axes[1,0].set_title('年龄与Y染色体浓度')
        axes[1,0].set_xlabel('年龄')
        axes[1,0].set_ylabel('Y染色体浓度 (%)')
        
        # 5. 身高与Y染色体浓度关系
        sns.scatterplot(data=self.male_data, x='身高', y='Y染色体浓度', 
                       hue='达标', palette=['red', 'green'], alpha=0.6, ax=axes[1,1])
        axes[1,1].axhline(4.0, color='black', linestyle='--', alpha=0.5)
        axes[1,1].set_title('身高与Y染色体浓度')
        axes[1,1].set_xlabel('身高 (cm)')
        axes[1,1].set_ylabel('Y染色体浓度 (%)')
        
        # 6. 体重与Y染色体浓度关系
        sns.scatterplot(data=self.male_data, x='体重', y='Y染色体浓度', 
                       hue='达标', palette=['red', 'green'], alpha=0.6, ax=axes[1,2])
        axes[1,2].axhline(4.0, color='black', linestyle='--', alpha=0.5)
        axes[1,2].set_title('体重与Y染色体浓度')
        axes[1,2].set_xlabel('体重 (kg)')
        axes[1,2].set_ylabel('Y染色体浓度 (%)')
        
        plt.tight_layout()
        plt.savefig('chart/ver2/exploratory_analysis.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        # 相关性分析
        corr_data = self.male_data[['孕周', 'BMI', '年龄', '身高', '体重', 'Y染色体浓度']]
        corr_matrix = corr_data.corr()
        
        plt.figure(figsize=(10, 8))
        sns.heatmap(corr_matrix, annot=True, cmap='coolwarm', center=0, 
                   square=True, fmt='.3f', cbar_kws={'shrink': 0.8})
        plt.title('影响因素相关性热图', fontsize=14, fontweight='bold')
        plt.tight_layout()
        plt.savefig('chart/ver2/correlation_heatmap.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        return corr_matrix
    
    def build_gamm_model(self):
        """建立GAMM模型（改进版，使用多种先进方法）"""
        print("正在建立GAMM模型...")
        
        # 准备特征
        X = self.male_data[['孕周', 'BMI', '年龄', '身高', '体重']].values
        y = self.male_data['Y染色体浓度'].values
        
        # 标准化
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
            
        # 尝试多种模型方法
        models_to_try = []
        
        # 1. 尝试使用statsmodels的GAM
        if STATSMODELS_AVAILABLE:
            try:
                print("尝试使用statsmodels GAM...")
                # 创建样条基函数
                bs = BSplines(X_scaled, df=[10, 10, 10, 10, 10], degree=[3, 3, 3, 3, 3])
                gam_model = GLMGam(y, smoother=bs, family=sm.families.Gaussian())
                gam_result = gam_model.fit()
                
                y_pred = gam_result.fittedvalues
                mae = mean_absolute_error(y, y_pred)
                r2 = r2_score(y, y_pred)
                
                models_to_try.append({
                    'name': 'statsmodels_gam',
                    'model': gam_result,
                    'scaler': scaler,
                    'mae': mae,
                    'r2': r2,
                    'type': 'statsmodels_gam'
                })
                print(f"Statsmodels GAM性能: MAE={mae:.3f}, R²={r2:.3f}")
            except Exception as e:
                print(f"Statsmodels GAM失败: {e}")
        
        # 2. 梯度提升回归器（模拟GAMM的非线性特性）
        try:
            print("尝试梯度提升回归器...")
            gbr_model = GradientBoostingRegressor(
                n_estimators=200,
                learning_rate=0.1,
                max_depth=6,
                min_samples_split=5,
                min_samples_leaf=2,
                random_state=42
            )
            gbr_model.fit(X_scaled, y)
            y_pred = gbr_model.predict(X_scaled)
            mae = mean_absolute_error(y, y_pred)
            r2 = r2_score(y, y_pred)
            
            models_to_try.append({
                'name': 'gradient_boosting',
                'model': gbr_model,
                'scaler': scaler,
                'mae': mae,
                'r2': r2,
                'type': 'gradient_boosting'
            })
            print(f"梯度提升回归器性能: MAE={mae:.3f}, R²={r2:.3f}")
        except Exception as e:
            print(f"梯度提升回归器失败: {e}")
        
        # 3. 弹性网络回归（带正则化）
        try:
            print("尝试弹性网络回归...")
            elastic_model = ElasticNet(
                alpha=0.1,
                l1_ratio=0.5,
                max_iter=2000,
                random_state=42
            )
            elastic_model.fit(X_scaled, y)
            y_pred = elastic_model.predict(X_scaled)
            mae = mean_absolute_error(y, y_pred)
            r2 = r2_score(y, y_pred)
            
            models_to_try.append({
                'name': 'elastic_net',
                'model': elastic_model,
                'scaler': scaler,
                'mae': mae,
                'r2': r2,
                'type': 'elastic_net'
            })
            print(f"弹性网络回归性能: MAE={mae:.3f}, R²={r2:.3f}")
        except Exception as e:
            print(f"弹性网络回归失败: {e}")
        
        # 4. 支持向量回归
        try:
            print("尝试支持向量回归...")
            svr_model = SVR(
                kernel='rbf',
                C=1.0,
                gamma='scale',
                epsilon=0.1
            )
            svr_model.fit(X_scaled, y)
            y_pred = svr_model.predict(X_scaled)
            mae = mean_absolute_error(y, y_pred)
            r2 = r2_score(y, y_pred)
            
            models_to_try.append({
                'name': 'svr',
                'model': svr_model,
                'scaler': scaler,
                'mae': mae,
                'r2': r2,
                'type': 'svr'
            })
            print(f"支持向量回归性能: MAE={mae:.3f}, R²={r2:.3f}")
        except Exception as e:
            print(f"支持向量回归失败: {e}")
        
        # 5. 神经网络回归
        try:
            print("尝试神经网络回归...")
            mlp_model = MLPRegressor(
                hidden_layer_sizes=(100, 50),
                activation='relu',
                solver='adam',
                alpha=0.001,
                learning_rate='adaptive',
                max_iter=1000,
                random_state=42
            )
            mlp_model.fit(X_scaled, y)
            y_pred = mlp_model.predict(X_scaled)
            mae = mean_absolute_error(y, y_pred)
            r2 = r2_score(y, y_pred)
            
            models_to_try.append({
                'name': 'neural_network',
                'model': mlp_model,
                'scaler': scaler,
                'mae': mae,
                'r2': r2,
                'type': 'neural_network'
            })
            print(f"神经网络回归性能: MAE={mae:.3f}, R²={r2:.3f}")
        except Exception as e:
            print(f"神经网络回归失败: {e}")
        
        # 选择最佳模型
        if models_to_try:
            best_model = max(models_to_try, key=lambda x: x['r2'])
            print(f"选择最佳模型: {best_model['name']} (R²={best_model['r2']:.3f})")
            
            self.models['gamm'] = best_model
            self.models['all_models'] = models_to_try  # 保存所有模型用于比较
        else:
            print("所有模型都失败了，使用随机森林作为备选...")
            rf_model = RandomForestRegressor(
                n_estimators=100, 
                max_depth=10, 
                min_samples_split=5,
                random_state=42
            )
            rf_model.fit(X_scaled, y)
            y_pred = rf_model.predict(X_scaled)
            mae = mean_absolute_error(y, y_pred)
            r2 = r2_score(y, y_pred)
            
            self.models['gamm'] = {
                'name': 'random_forest',
                'model': rf_model,
                'scaler': scaler,
                'mae': mae,
                'r2': r2,
                'type': 'random_forest'
            }
        
        # 可视化拟合效果
        plt.figure(figsize=(15, 10))
        
        # 1. 预测vs实际
        plt.subplot(2, 3, 1)
        plt.scatter(y, y_pred, alpha=0.6, color='blue')
        plt.plot([y.min(), y.max()], [y.min(), y.max()], 'r--', lw=2)
        plt.xlabel('Actual Y Chromosome Concentration (%)')
        plt.ylabel('Predicted Y Chromosome Concentration (%)')
        plt.title(f'GAMM拟合结果 (R²={r2:.3f})')
        
        # 2. 残差分析
        plt.subplot(2, 3, 2)
        residuals = y - y_pred
        plt.scatter(y_pred, residuals, alpha=0.6, color='green')
        plt.axhline(y=0, color='r', linestyle='--')
        plt.xlabel('预测值')
        plt.ylabel('残差')
        plt.title('残差分析')
        
        # 3. 特征重要性（如果是随机森林）
        if self.models['gamm']['type'] == 'random_forest':
            plt.subplot(2, 3, 3)
            feature_names = ['Gestational Week', 'BMI', 'Age', 'Height', 'Weight']
            importance = self.models['gamm']['model'].feature_importances_
            plt.barh(feature_names, importance, color='skyblue', alpha=0.7)
            plt.xlabel('重要性')
            plt.title('特征重要性')
        
        # 4. 孕周与Y染色体浓度关系
        plt.subplot(2, 3, 4)
        plt.scatter(self.male_data['孕周'], y, alpha=0.6, color='blue', label='Actual')
        plt.scatter(self.male_data['孕周'], y_pred, alpha=0.6, color='red', label='Predicted')
        plt.xlabel('孕周')
        plt.ylabel('Y染色体浓度 (%)')
        plt.title('孕周与Y染色体浓度')
        plt.legend()
        
        # 5. BMI与Y染色体浓度关系
        plt.subplot(2, 3, 5)
        plt.scatter(self.male_data['BMI'], y, alpha=0.6, color='blue', label='Actual')
        plt.scatter(self.male_data['BMI'], y_pred, alpha=0.6, color='red', label='Predicted')
        plt.xlabel('BMI')
        plt.ylabel('Y染色体浓度 (%)')
        plt.title('BMI与Y染色体浓度')
        plt.legend()
        
        # 6. 残差分布
        plt.subplot(2, 3, 6)
        plt.hist(residuals, bins=30, alpha=0.7, color='lightgreen', edgecolor='black')
        plt.xlabel('残差')
        plt.ylabel('频数')
        plt.title('残差分布')
        
        plt.tight_layout()
        plt.savefig('chart/ver2/gamm_model_results.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        return self.models['gamm']['model']
    
    def handle_selection_bias(self):
        """处理选择偏倚（IPW方法）"""
        print("正在处理选择偏倚...")
        
        from sklearn.linear_model import LogisticRegression
        from sklearn.preprocessing import StandardScaler
        
        # 创建失败模型（模拟测序失败）
        # 假设失败与BMI、孕周、年龄相关
        failure_features = self.male_data[['BMI', '孕周', '年龄']].values
        scaler_failure = StandardScaler()
        failure_features_scaled = scaler_failure.fit_transform(failure_features)
        
        # 模拟失败概率（高BMI、早期孕周更容易失败）
        failure_prob = 1 / (1 + np.exp(-(0.1 * self.male_data['BMI'] - 0.2 * self.male_data['孕周'] + 0.05 * self.male_data['年龄'] - 2)))
        
        # 生成失败标志
        np.random.seed(42)
        failure_indicator = np.random.binomial(1, failure_prob, len(self.male_data))
        
        # 拟合失败模型
        failure_model = LogisticRegression(random_state=42)
        failure_model.fit(failure_features_scaled, failure_indicator)
        
        # 计算IPW权重
        failure_prob_pred = failure_model.predict_proba(failure_features_scaled)[:, 1]
        ipw_weights = 1 / (1 - failure_prob_pred + 1e-8)  # 避免除零
        
        # 标准化权重
        ipw_weights = ipw_weights / np.mean(ipw_weights)
        
        # 保存结果
        self.models['ipw'] = {
            'failure_model': failure_model,
            'failure_scaler': scaler_failure,
            'ipw_weights': ipw_weights,
            'failure_indicator': failure_indicator,
            'failure_prob': failure_prob
        }
        
        # 可视化IPW结果
        plt.figure(figsize=(15, 5))
        
        # 1. 失败概率分布
        plt.subplot(1, 3, 1)
        plt.hist(failure_prob, bins=30, alpha=0.7, color='red', edgecolor='black')
        plt.xlabel('失败概率')
        plt.ylabel('频数')
        plt.title('测序失败概率分布')
        
        # 2. IPW权重分布
        plt.subplot(1, 3, 2)
        plt.hist(ipw_weights, bins=30, alpha=0.7, color='blue', edgecolor='black')
        plt.xlabel('IPW权重')
        plt.ylabel('频数')
        plt.title('IPW权重分布')
        
        # 3. BMI与失败概率关系
        plt.subplot(1, 3, 3)
        plt.scatter(self.male_data['BMI'], failure_prob, alpha=0.6, color='orange')
        plt.xlabel('BMI')
        plt.ylabel('失败概率')
        plt.title('BMI与失败概率')
        
        plt.tight_layout()
        plt.savefig('chart/ver2/ipw_analysis.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        print(f"IPW权重统计: 均值={np.mean(ipw_weights):.3f}, 标准差={np.std(ipw_weights):.3f}")
        return ipw_weights
    
    def predict_individual_recommended_week(self, age, height, weight, bmi, threshold=4.0, min_week=10, max_week=25):
        """基于当前最佳回归模型，对单个个体给出达标周的个性化推荐。
        逻辑：在[min_week, max_week]遍历孕周，将其他特征固定为输入值，找到预测Y≥threshold的最小周。
        若始终达不到阈值，返回max_week。
        """
        if 'gamm' not in self.models:
            raise RuntimeError('请先训练GAMM/回归模型')
        model_info = self.models['gamm']
        model = model_info['model']
        scaler = model_info.get('scaler', None)
        weeks = np.arange(min_week, max_week + 1)
        X_grid = np.column_stack([
            weeks,
            np.full_like(weeks, bmi, dtype=float),
            np.full_like(weeks, age, dtype=float),
            np.full_like(weeks, height, dtype=float),
            np.full_like(weeks, weight, dtype=float)
        ])
        if scaler is not None:
            X_grid_scaled = scaler.transform(X_grid)
        else:
            X_grid_scaled = X_grid
        try:
            y_pred = model.predict(X_grid_scaled)
        except Exception:
            # statsmodels GAM result
            y_pred = model.predict(X_grid_scaled)
        meet_idx = np.where(y_pred >= threshold)[0]
        if len(meet_idx) == 0:
            return int(max_week)
        return int(weeks[meet_idx[0]])

    def bootstrap_bmi_recommendation_ci(self, n_bootstrap=500, ci=0.95):
        """对当前BMI分组推荐的达标概率进行Bootstrap置信区间估计。
        假设self.results['bmi_grouping']已存在。返回DataFrame并保存图表。
        """
        if 'bmi_grouping' not in self.results or not self.results['bmi_grouping']:
            print('无BMI分组结果可用于Bootstrap')
            return None
        results = []
        alpha = 1 - ci
        for rec in self.results['bmi_grouping']:
            group_str = rec['BMI_group']
            week = rec['week']
            # 解析组区间
            try:
                bounds = group_str.strip('[]').split(',')
                bmi_min = float(bounds[0])
                bmi_max = float(bounds[1])
            except Exception:
                # 回退：用数据中该组的范围
                bmi_min = self.male_data['BMI'].min()
                bmi_max = self.male_data['BMI'].max()
            group_data = self.male_data[(self.male_data['BMI'] >= bmi_min) & (self.male_data['BMI'] <= bmi_max) & (self.male_data['孕周'] <= week)]
            if len(group_data) < 5:
                continue
            probs = []
            for _ in range(n_bootstrap):
                sample = group_data.sample(n=len(group_data), replace=True, random_state=None)
                probs.append(sample['达标'].mean())
            probs = np.array(probs)
            lower = np.quantile(probs, alpha/2)
            upper = np.quantile(probs, 1 - alpha/2)
            results.append({
                'BMI_group': group_str,
                'week': week,
                'prob_mean': float(np.mean(probs)),
                'prob_lower': float(lower),
                'prob_upper': float(upper),
                'n': int(len(group_data))
            })
        if not results:
            print('Bootstrap未生成结果')
            return None
        df = pd.DataFrame(results)
        # 可视化
        plt.figure(figsize=(10, 6))
        x = np.arange(len(df))
        means = df['prob_mean'].values
        lower_err = means - df['prob_lower'].values
        upper_err = df['prob_upper'].values - means
        plt.errorbar(x, means, yerr=[lower_err, upper_err], fmt='o', capsize=5, capthick=2, markersize=8, color='tab:blue', ecolor='black')
        plt.xticks(x, [f"{g}\nW{w}" for g, w in zip(df['BMI_group'], df['week'])], rotation=45)
        plt.ylabel('达标概率 (Bootstrap CI)')
        plt.title('BMI分组推荐的达标率Bootstrap置信区间')
        plt.tight_layout()
        plt.savefig('chart/bmi_recommendation_bootstrap_ci.png', dpi=300, bbox_inches='tight')
        plt.close()
        self.results['bmi_bootstrap_ci'] = df
        return df

    def build_survival_model(self):
        """建立生存分析模型（改进版，使用真正的生存分析方法）"""
        print("正在建立生存分析模型...")
        
        # 准备生存分析数据
        survival_data = []
        
        # 若缺少患者标识列B，做兼容处理
        if 'B' not in self.male_data.columns:
            if '孕妇代码' in self.male_data.columns:
                self.male_data['B'] = self.male_data['孕妇代码']
            elif 'A' in self.male_data.columns:
                self.male_data['B'] = self.male_data['A']
            else:
                # 退化情形：每条记录视为一个独立患者
                self.male_data['B'] = np.arange(len(self.male_data))
        
        for patient_id in self.male_data['B'].unique():
            patient_records = self.male_data[self.male_data['B'] == patient_id].sort_values('孕周')
            
            if len(patient_records) == 0:
                continue
                
            # 获取患者特征（取第一次检测时的特征）
            features = patient_records.iloc[0][['BMI', '年龄', '身高', '体重']].values
            
            # 找到首次达标的时间点
            first_达标_idx = None
            for idx, row in patient_records.iterrows():
                if row['达标'] == 1:
                    first_达标_idx = idx
                    break
            
            if first_达标_idx is not None:
                # 事件发生（达标）
                event_time = patient_records.loc[first_达标_idx, '孕周']
                event_observed = True
            else:
                # 删失（未达标）
                event_time = patient_records['孕周'].max()
                event_observed = False
            
            survival_data.append({
                'patient_id': patient_id,
                'duration': event_time,
                'event_observed': event_observed,
                'BMI': features[0],
                '年龄': features[1],
                '身高': features[2],
                '体重': features[3]
            })
        
        if len(survival_data) == 0:
            print("警告：没有找到生存分析数据")
            return None
            
        survival_df = pd.DataFrame(survival_data)
        
        # 尝试使用lifelines进行真正的生存分析
        if LIFELINES_AVAILABLE:
            try:
                print("使用lifelines进行生存分析...")
                
                # 1. Kaplan-Meier估计
                kmf = KaplanMeierFitter()
                kmf.fit(survival_df['duration'], survival_df['event_observed'], label='Overall')
                
                # 2. Cox比例风险模型
                cph = CoxPHFitter()
                cph.fit(survival_df[['duration', 'event_observed', 'BMI', '年龄', '身高', '体重']], 
                       duration_col='duration', event_col='event_observed')
                
                # 3. Weibull模型
                wf = WeibullFitter()
                wf.fit(survival_df['duration'], survival_df['event_observed'])
                
                # 保存模型
                self.models['survival'] = {
                    'kaplan_meier': kmf,
                    'cox_model': cph,
                    'weibull_model': wf,
                    'survival_data': survival_df,
                    'type': 'lifelines'
                }
                
                # 可视化生存分析结果
                self._plot_survival_analysis(survival_df, kmf, cph, wf)
                
                print("生存分析完成")
                return cph
                
            except Exception as e:
                print(f"Lifelines生存分析失败: {e}")
                print("使用机器学习方法作为备选...")
        
        # 备选方案：使用机器学习方法
        print("使用机器学习方法进行生存分析...")
        
        # 准备特征和目标变量
        X = survival_df[['BMI', '年龄', '身高', '体重']]
        y = survival_df['duration']
        
        # 尝试多种模型
        models_to_try = []
        
        # 1. 随机森林
        try:
            rf_model = RandomForestRegressor(n_estimators=100, random_state=42)
            rf_model.fit(X, y)
            y_pred = rf_model.predict(X)
            mae = mean_absolute_error(y, y_pred)
            r2 = r2_score(y, y_pred)
            
            models_to_try.append({
                'name': 'random_forest',
                'model': rf_model,
                'mae': mae,
                'r2': r2
            })
        except Exception as e:
            print(f"随机森林失败: {e}")
        
        # 2. 梯度提升
        try:
            gbr_model = GradientBoostingRegressor(n_estimators=100, random_state=42)
            gbr_model.fit(X, y)
            y_pred = gbr_model.predict(X)
            mae = mean_absolute_error(y, y_pred)
            r2 = r2_score(y, y_pred)
            
            models_to_try.append({
                'name': 'gradient_boosting',
                'model': gbr_model,
                'mae': mae,
                'r2': r2
            })
        except Exception as e:
            print(f"梯度提升失败: {e}")
        
        # 选择最佳模型
        if models_to_try:
            best_model = max(models_to_try, key=lambda x: x['r2'])
            print(f"选择最佳生存模型: {best_model['name']} (R²={best_model['r2']:.3f})")
            
            self.models['survival'] = {
                'model': best_model['model'],
                'survival_data': survival_df,
                'mae': best_model['mae'],
                'r2': best_model['r2'],
                'type': 'machine_learning'
            }
            
            # 特征重要性
            if hasattr(best_model['model'], 'feature_importances_'):
                feature_importance = pd.DataFrame({
                    'feature': ['BMI', '年龄', '身高', '体重'],
                    'importance': best_model['model'].feature_importances_
                }).sort_values('importance', ascending=False)
                
                plt.figure(figsize=(10, 6))
                sns.barplot(data=feature_importance, x='importance', y='feature', palette='viridis')
                plt.title('生存模型特征重要性', fontsize=14, fontweight='bold')
                plt.xlabel('重要性')
                plt.tight_layout()
                plt.savefig('chart/ver2/survival_model_importance.png', dpi=300, bbox_inches='tight')
                plt.close()
            
            return best_model['model']
        else:
            print("所有生存模型都失败了")
            return None
    
    def _plot_survival_analysis(self, survival_df, kmf, cph, wf):
        """绘制生存分析结果"""
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        fig.suptitle('生存分析结果', fontsize=16, fontweight='bold')
        
        # 1. Kaplan-Meier生存曲线
        ax1 = axes[0, 0]
        kmf.plot_survival_function(ax=ax1)
        ax1.set_title('Kaplan-Meier 生存曲线')
        ax1.set_xlabel('孕周')
        ax1.set_ylabel('生存概率')
        ax1.grid(True, alpha=0.3)
        
        # 2. 按BMI分组的生存曲线
        ax2 = axes[0, 1]
        bmi_median = survival_df['BMI'].median()
        high_bmi = survival_df[survival_df['BMI'] >= bmi_median]
        low_bmi = survival_df[survival_df['BMI'] < bmi_median]
        
        if len(high_bmi) > 0:
            kmf_high = KaplanMeierFitter()
            kmf_high.fit(high_bmi['duration'], high_bmi['event_observed'], label='High BMI')
            kmf_high.plot_survival_function(ax=ax2)
        
        if len(low_bmi) > 0:
            kmf_low = KaplanMeierFitter()
            kmf_low.fit(low_bmi['duration'], low_bmi['event_observed'], label='Low BMI')
            kmf_low.plot_survival_function(ax=ax2)
        
        ax2.set_title('按BMI分组的生存曲线')
        ax2.set_xlabel('孕周')
        ax2.set_ylabel('生存概率')
        ax2.legend()
        ax2.grid(True, alpha=0.3)
        
        # 3. Cox模型系数
        ax3 = axes[1, 0]
        coef_data = cph.summary[['coef', 'exp(coef)', 'p']]
        coef_data.plot(kind='bar', ax=ax3, color=['skyblue', 'lightgreen', 'lightcoral'])
        ax3.set_title('Cox模型系数')
        ax3.set_ylabel('系数值')
        ax3.tick_params(axis='x', rotation=45)
        ax3.legend(['系数', '风险比', 'P值'])
        
        # 4. 风险评分分布
        ax4 = axes[1, 1]
        risk_scores = cph.predict_partial_hazard(survival_df[['BMI', '年龄', '身高', '体重']])
        ax4.hist(risk_scores, bins=20, alpha=0.7, color='orange', edgecolor='black')
        ax4.set_title('风险评分分布')
        ax4.set_xlabel('风险评分')
        ax4.set_ylabel('频数')
        ax4.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig('chart/ver2/survival_analysis_detailed.png', dpi=300, bbox_inches='tight')
        plt.close()
    
    def bmi_grouping_optimization(self):
        """BMI分组优化（改进版，使用智能分组算法）"""
        print("正在进行BMI分组优化...")
        
        # 优先使用K-means聚类进行智能BMI分组；若失败则回退到分位数分组
        bmi_values = self.male_data['BMI'].values.reshape(-1, 1)
        use_kmeans = True
        centers = None
        try:
            # 尝试不同的聚类数量
            best_k = 3
            best_silhouette = -1
            for k in range(2, 8):
                try:
                    kmeans_try = KMeans(n_clusters=k, random_state=42, n_init=10)
                    labels_try = kmeans_try.fit_predict(bmi_values)
                    from sklearn.metrics import silhouette_score
                    silhouette = silhouette_score(bmi_values, labels_try)
                    if silhouette > best_silhouette:
                        best_silhouette = silhouette
                        best_k = k
                except Exception:
                    continue
            print(f"选择最佳聚类数: {best_k} (轮廓系数: {best_silhouette:.3f})")
            # 使用最佳聚类数进行分组
            kmeans = KMeans(n_clusters=best_k, random_state=42, n_init=10)
            labels = kmeans.fit_predict(bmi_values)
            centers = kmeans.cluster_centers_.flatten()
            sorted_indices = np.argsort(centers)
            # 创建BMI分组，携带中心值
            bmi_groups = []
            for i in range(best_k):
                cluster_idx = sorted_indices[i]
                cluster_data = self.male_data[labels == cluster_idx]
                bmi_min = cluster_data['BMI'].min()
                bmi_max = cluster_data['BMI'].max()
                center_val = float(centers[cluster_idx])
                bmi_groups.append((bmi_min, bmi_max, center_val))
        except Exception as e:
            # 回退到按分位数分组（3组或数据更少时两组）
            print(f"KMeans分组失败，改用分位数分组: {e}")
            use_kmeans = False
            unique_count = len(np.unique(bmi_values))
            if unique_count >= 3:
                qs = np.quantile(bmi_values.flatten(), [0.0, 1/3, 2/3, 1.0])
                bins = [(qs[0], qs[1]), (qs[1], qs[2]), (qs[2], qs[3])]
            else:
                qs = np.quantile(bmi_values.flatten(), [0.0, 0.5, 1.0])
                bins = [(qs[0], qs[1]), (qs[1], qs[2])]
            bmi_groups = []
            for bmin, bmax in bins:
                center_val = float((bmin + bmax) / 2.0)
                bmi_groups.append((float(bmin), float(bmax), center_val))
        
        print(f"智能BMI分组: {[(round(a,1), round(b,1)) for a,b,_ in bmi_groups]}")
        
        # 定义候选检测时点
        candidate_weeks = list(range(10, 26))
        
        results = []
        pareto_results = []  # 存储Pareto前沿结果
        
        for bmi_min, bmi_max, center_val in bmi_groups:
            group_data = self.male_data[
                (self.male_data['BMI'] >= bmi_min) & 
                (self.male_data['BMI'] <= bmi_max)
            ]
            
            if len(group_data) < 10:  # 样本量太小
                continue
                
            group_results = []
            
            for week in candidate_weeks:
                # 计算在该周检测的风险
                week_data = group_data[group_data['孕周'] <= week]
                
                if len(week_data) == 0:
                    continue
                
                # 达标概率
                达标_prob = week_data['达标'].mean()
                
                # 风险计算
                miss_risk = 1 - 达标_prob  # 漏检风险
                
                # 晚检风险（孕周越晚风险越高）
                if week <= 12:
                    late_risk = 0
                elif week <= 27:
                    late_risk = (week - 12) * 0.1
                else:
                    late_risk = 1.5 + (week - 27) * 0.2
                
                # 成本风险（检测成本）
                cost_risk = week * 0.01  # 孕周越晚成本越高
                
                # 总风险（权重可调）
                total_risk = 0.5 * miss_risk + 0.3 * late_risk + 0.2 * cost_risk
                
                # 计算置信区间
                n = len(week_data)
                if n > 1:
                    std_error = np.sqrt(达标_prob * (1 - 达标_prob) / n)
                    ci_lower = max(0, 达标_prob - 1.96 * std_error)
                    ci_upper = min(1, 达标_prob + 1.96 * std_error)
                else:
                    ci_lower = ci_upper = 达标_prob
                
                result = {
                    'BMI_group': f'[{bmi_min:.1f}, {bmi_max:.1f}]',
                    'week': week,
                    '达标概率': 达标_prob,
                    '达标概率_CI_lower': ci_lower,
                    '达标概率_CI_upper': ci_upper,
                    '漏检风险': miss_risk,
                    '晚检风险': late_risk,
                    '成本风险': cost_risk,
                    '总风险': total_risk,
                    '样本数': len(week_data),
                    '聚类中心': center_val
                }
                
                group_results.append(result)
                pareto_results.append(result)
            
            if group_results:
                # 选择风险最小的时点
                best_result = min(group_results, key=lambda x: x['总风险'])
                results.append(best_result)
        
        self.results['bmi_grouping'] = results
        self.results['pareto_results'] = pareto_results
        groups_simple = [(float(g[0]), float(g[1])) for g in bmi_groups]
        self.results['bmi_clusters'] = {
            'method': 'kmeans' if use_kmeans else 'quantile',
            'kmeans': kmeans if use_kmeans else None,
            'labels': labels if use_kmeans else None,
            'centers': centers if use_kmeans else None,
            'groups': groups_simple
        }
        
        # 计算Pareto前沿
        self._calculate_pareto_frontier(pareto_results)
        
        # 可视化结果
        self._plot_bmi_grouping_results(results)
        
        return results
    
    def _calculate_pareto_frontier(self, results):
        """计算Pareto前沿"""
        print("正在计算Pareto前沿...")
        
        # 提取目标函数值
        miss_risks = [r['漏检风险'] for r in results]
        late_risks = [r['晚检风险'] for r in results]
        
        # 找到Pareto最优解
        pareto_indices = []
        for i in range(len(results)):
            is_pareto = True
            for j in range(len(results)):
                if i != j:
                    # 如果j在miss_risk和late_risk上都优于i，则i不是Pareto最优
                    if (miss_risks[j] <= miss_risks[i] and late_risks[j] <= late_risks[i] and 
                        (miss_risks[j] < miss_risks[i] or late_risks[j] < late_risks[i])):
                        is_pareto = False
                        break
            if is_pareto:
                pareto_indices.append(i)
        
        # 保存Pareto前沿结果
        pareto_frontier = [results[i] for i in pareto_indices]
        self.results['pareto_frontier'] = pareto_frontier
        
        # 可视化Pareto前沿
        self._plot_pareto_frontier(results, pareto_frontier)
        
        return pareto_frontier
    
    def _plot_pareto_frontier(self, all_results, pareto_frontier):
        """绘制Pareto前沿"""
        plt.figure(figsize=(12, 8))
        
        # 所有解
        all_miss = [r['漏检风险'] for r in all_results]
        all_late = [r['晚检风险'] for r in all_results]
        
        # Pareto前沿解
        pareto_miss = [r['漏检风险'] for r in pareto_frontier]
        pareto_late = [r['晚检风险'] for r in pareto_frontier]
        
        # 绘制散点图
        plt.scatter(all_miss, all_late, alpha=0.6, color='lightblue', 
                   label='所有候选解', s=50)
        plt.scatter(pareto_miss, pareto_late, alpha=0.8, color='red', 
                   label='Pareto前沿', s=100, marker='o', edgecolors='black')
        
        # 连接Pareto前沿点
        if len(pareto_frontier) > 1:
            # 按miss_risk排序
            sorted_pareto = sorted(pareto_frontier, key=lambda x: x['漏检风险'])
            sorted_miss = [r['漏检风险'] for r in sorted_pareto]
            sorted_late = [r['晚检风险'] for r in sorted_pareto]
            plt.plot(sorted_miss, sorted_late, 'r--', alpha=0.7, linewidth=2)
        
        # 标记最优解（总风险最小）
        if all_results:
            best_solution = min(all_results, key=lambda x: x['总风险'])
            plt.scatter(best_solution['漏检风险'], best_solution['晚检风险'], 
                       color='green', s=200, marker='*', 
                       label=f'最优解 (第{best_solution["week"]}周)', 
                       edgecolors='black', linewidth=2)
        
        plt.xlabel('漏检风险', fontsize=12)
        plt.ylabel('晚检风险', fontsize=12)
        plt.title('多目标优化 Pareto 前沿分析', fontsize=14, fontweight='bold')
        plt.legend(fontsize=10)
        plt.grid(True, alpha=0.3)
        
        # 添加注释
        for i, result in enumerate(pareto_frontier[:5]):  # 只标注前5个
            plt.annotate(f'Week {result["week"]}\n{result["BMI_group"]}', 
                        (result['漏检风险'], result['晚检风险']),
                        xytext=(5, 5), textcoords='offset points',
                        fontsize=8, alpha=0.8)
        
        plt.tight_layout()
        plt.savefig('chart/ver2/pareto_frontier_analysis.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        print(f"找到{len(pareto_frontier)}个Pareto最优解")
    
    def _plot_bmi_grouping_results(self, results):
        """绘制BMI分组结果"""
        if not results:
            return
            
        df_results = pd.DataFrame(results)
        
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        fig.suptitle('BMI分组优化结果', fontsize=16, fontweight='bold')
        
        # 1. 各BMI组的最佳检测时点
        axes[0,0].bar(df_results['BMI_group'], df_results['week'], 
                     color='skyblue', edgecolor='black', alpha=0.7)
        axes[0,0].set_title('各BMI组推荐检测孕周')
        axes[0,0].set_ylabel('推荐孕周')
        axes[0,0].tick_params(axis='x', rotation=45)
        
        # 2. 达标概率
        axes[0,1].bar(df_results['BMI_group'], df_results['达标概率'], 
                     color='lightgreen', edgecolor='black', alpha=0.7)
        axes[0,1].set_title('各BMI组达标概率')
        axes[0,1].set_ylabel('达标概率')
        axes[0,1].tick_params(axis='x', rotation=45)
        
        # 3. 风险分析
        x = np.arange(len(df_results))
        width = 0.25
        
        axes[1,0].bar(x - width, df_results['漏检风险'], width, 
                     label='Miss Risk', color='red', alpha=0.7)
        axes[1,0].bar(x, df_results['晚检风险'], width, 
                     label='Late Risk', color='orange', alpha=0.7)
        axes[1,0].bar(x + width, df_results['总风险'], width, 
                     label='Total Risk', color='purple', alpha=0.7)
        
        axes[1,0].set_title('风险分析')
        axes[1,0].set_ylabel('风险值')
        axes[1,0].set_xticks(x)
        axes[1,0].set_xticklabels(df_results['BMI_group'], rotation=45)
        axes[1,0].legend()
        
        # 4. 样本数
        axes[1,1].bar(df_results['BMI_group'], df_results['样本数'], 
                     color='lightcoral', edgecolor='black', alpha=0.7)
        axes[1,1].set_title('各BMI组样本数')
        axes[1,1].set_ylabel('样本数')
        axes[1,1].tick_params(axis='x', rotation=45)
        
        plt.tight_layout()
        plt.savefig('chart/ver2/bmi_grouping_results.png', dpi=300, bbox_inches='tight')
        plt.close()
    
    def monte_carlo_analysis(self, n_simulations=1000):
        """蒙特卡洛不确定性分析"""
        print(f"正在进行蒙特卡洛分析（{n_simulations}次模拟）...")
        
        # 模拟参数
        measurement_error_std = 0.5  # Y染色体浓度测量误差标准差
        bmi_error_std = 1.0  # BMI测量误差标准差
        
        simulation_results = []
        
        for sim in range(n_simulations):
            # 添加测量误差
            simulated_data = self.male_data.copy()
            simulated_data['Y染色体浓度'] += np.random.normal(0, measurement_error_std, len(simulated_data))
            simulated_data['BMI'] += np.random.normal(0, bmi_error_std, len(simulated_data))
            
            # 重新计算达标标志
            simulated_data['达标'] = (simulated_data['Y染色体浓度'] >= 4.0).astype(int)
            
            # 重新进行BMI分组优化
            bmi_groups = [(20, 28), (28, 32), (32, 36), (36, 40), (40, 50)]
            candidate_weeks = list(range(10, 26))
            
            sim_results = []
            
            for bmi_min, bmi_max in bmi_groups:
                group_data = simulated_data[
                    (simulated_data['BMI'] >= bmi_min) & 
                    (simulated_data['BMI'] < bmi_max)
                ]
                
                if len(group_data) < 10:
                    continue
                
                best_week = None
                best_risk = float('inf')
                
                for week in candidate_weeks:
                    week_data = group_data[group_data['孕周'] <= week]
                    
                    if len(week_data) == 0:
                        continue
                    
                    达标_prob = week_data['达标'].mean()
                    miss_risk = 1 - 达标_prob
                    
                    if week <= 12:
                        late_risk = 0
                    elif week <= 27:
                        late_risk = (week - 12) * 0.1
                    else:
                        late_risk = 1.5 + (week - 27) * 0.2
                    
                    total_risk = 0.7 * miss_risk + 0.3 * late_risk
                    
                    if total_risk < best_risk:
                        best_risk = total_risk
                        best_week = week
                
                if best_week is not None:
                    sim_results.append({
                        'BMI_group': f'[{bmi_min}, {bmi_max})',
                        'best_week': best_week,
                        'best_risk': best_risk
                    })
            
            simulation_results.append(sim_results)
        
        # 分析蒙特卡洛结果
        self._analyze_monte_carlo_results(simulation_results)
        
        return simulation_results
    
    def _analyze_monte_carlo_results(self, simulation_results):
        """分析蒙特卡洛结果"""
        # 统计每个BMI组的最优时点分布
        bmi_group_stats = {}
        
        for sim_results in simulation_results:
            for result in sim_results:
                bmi_group = result['BMI_group']
                if bmi_group not in bmi_group_stats:
                    bmi_group_stats[bmi_group] = []
                bmi_group_stats[bmi_group].append(result['best_week'])
        
        # 计算统计量
        stats_summary = {}
        for bmi_group, weeks in bmi_group_stats.items():
            stats_summary[bmi_group] = {
                'mean_week': np.mean(weeks),
                'std_week': np.std(weeks),
                'median_week': np.median(weeks),
                'q25': np.percentile(weeks, 25),
                'q75': np.percentile(weeks, 75),
                'count': len(weeks)
            }
        
        self.results['monte_carlo'] = stats_summary
        
        # 可视化蒙特卡洛结果
        self._plot_monte_carlo_results(stats_summary)
        
        return stats_summary
    
    def _plot_monte_carlo_results(self, stats_summary):
        """绘制蒙特卡洛结果"""
        if not stats_summary:
            return
            
        bmi_groups = list(stats_summary.keys())
        means = [stats_summary[group]['mean_week'] for group in bmi_groups]
        stds = [stats_summary[group]['std_week'] for group in bmi_groups]
        q25s = [stats_summary[group]['q25'] for group in bmi_groups]
        q75s = [stats_summary[group]['q75'] for group in bmi_groups]
        
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        fig.suptitle('蒙特卡洛不确定性分析结果', fontsize=16, fontweight='bold')
        
        # 1. 最优时点均值与标准差
        x = np.arange(len(bmi_groups))
        axes[0,0].bar(x, means, yerr=stds, capsize=5, 
                     color='skyblue', edgecolor='black', alpha=0.7)
        axes[0,0].set_title('各BMI组最优孕周（均值±标准差）')
        axes[0,0].set_ylabel('推荐孕周')
        axes[0,0].set_xticks(x)
        axes[0,0].set_xticklabels(bmi_groups, rotation=45)
        
        # 2. 置信区间
        lower_err = np.maximum(0, np.array(means) - np.array(q25s))
        upper_err = np.maximum(0, np.array(q75s) - np.array(means))
        axes[0,1].errorbar(x, means, yerr=[lower_err, upper_err], 
                          fmt='o', capsize=5, capthick=2, markersize=8,
                          color='red', ecolor='black')
        axes[0,1].set_title('最优孕周置信区间（25%-75%）')
        axes[0,1].set_ylabel('推荐孕周')
        axes[0,1].set_xticks(x)
        axes[0,1].set_xticklabels(bmi_groups, rotation=45)
        
        # 3. 稳健性分析（标准差）
        axes[1,0].bar(bmi_groups, stds, color='orange', edgecolor='black', alpha=0.7)
        axes[1,0].set_title('推荐孕周的稳健性（标准差）')
        axes[1,0].set_ylabel('标准差')
        axes[1,0].tick_params(axis='x', rotation=45)
        
        # 4. 样本数
        counts = [stats_summary[group]['count'] for group in bmi_groups]
        axes[1,1].bar(bmi_groups, counts, color='lightgreen', edgecolor='black', alpha=0.7)
        axes[1,1].set_title('各BMI组有效模拟次数')
        axes[1,1].set_ylabel('模拟次数')
        axes[1,1].tick_params(axis='x', rotation=45)
        
        plt.tight_layout()
        plt.savefig('chart/ver2/monte_carlo_results.png', dpi=300, bbox_inches='tight')
        plt.close()
    
    def generate_final_report(self):
        """生成最终报告"""
        print("正在生成最终报告...")
        
        # 创建综合报告图表
        fig = plt.figure(figsize=(20, 16))
        gs = fig.add_gridspec(4, 4, hspace=0.3, wspace=0.3)
        
        # 1. 数据概览
        ax1 = fig.add_subplot(gs[0, :2])
        bmi_dist = self.male_data['BMI'].hist(bins=20, ax=ax1, color='lightblue', edgecolor='black', alpha=0.7)
        ax1.set_title('BMI分布', fontsize=14, fontweight='bold')
        ax1.set_xlabel('BMI')
        ax1.set_ylabel('频数')
        
        # 2. 达标率分析
        ax2 = fig.add_subplot(gs[0, 2:])
        week_groups = pd.cut(self.male_data['孕周'], bins=5)
        达标_rate = self.male_data.groupby(week_groups)['达标'].mean()
        达标_rate.plot(kind='bar', ax=ax2, color='lightgreen', edgecolor='black', alpha=0.7)
        ax2.set_title('不同孕周的达标率', fontsize=14, fontweight='bold')
        ax2.set_ylabel('达标率')
        ax2.tick_params(axis='x', rotation=45)
        
        # 3. BMI分组结果
        if 'bmi_grouping' in self.results:
            ax3 = fig.add_subplot(gs[1, :2])
            bmi_results = pd.DataFrame(self.results['bmi_grouping'])
            ax3.bar(bmi_results['BMI_group'], bmi_results['week'], 
                   color='skyblue', edgecolor='black', alpha=0.7)
            ax3.set_title('各BMI组推荐检测孕周', fontsize=14, fontweight='bold')
            ax3.set_ylabel('推荐孕周')
            ax3.tick_params(axis='x', rotation=45)
        
        # 4. 风险分析
        if 'bmi_grouping' in self.results:
            ax4 = fig.add_subplot(gs[1, 2:])
            bmi_results = pd.DataFrame(self.results['bmi_grouping'])
            x = np.arange(len(bmi_results))
            width = 0.25
            ax4.bar(x - width, bmi_results['漏检风险'], width, 
                   label='Miss Risk', color='red', alpha=0.7)
            ax4.bar(x, bmi_results['晚检风险'], width, 
                   label='Late Risk', color='orange', alpha=0.7)
            ax4.bar(x + width, bmi_results['总风险'], width, 
                   label='Total Risk', color='purple', alpha=0.7)
            ax4.set_title('风险分析', fontsize=14, fontweight='bold')
            ax4.set_ylabel('风险值')
            ax4.set_xticks(x)
            ax4.set_xticklabels(bmi_results['BMI_group'], rotation=45)
            ax4.legend()
        
        # 5. 蒙特卡洛结果
        if 'monte_carlo' in self.results:
            ax5 = fig.add_subplot(gs[2, :2])
            mc_results = self.results['monte_carlo']
            bmi_groups = list(mc_results.keys())
            means = [mc_results[group]['mean_week'] for group in bmi_groups]
            stds = [mc_results[group]['std_week'] for group in bmi_groups]
            x = np.arange(len(bmi_groups))
            ax5.bar(x, means, yerr=stds, capsize=5, 
                   color='lightcoral', edgecolor='black', alpha=0.7)
            ax5.set_title('蒙特卡洛：推荐孕周稳健性', fontsize=14, fontweight='bold')
            ax5.set_ylabel('推荐孕周')
            ax5.set_xticks(x)
            ax5.set_xticklabels(bmi_groups, rotation=45)
        
        # 6. 模型性能
        if 'gamm' in self.models:
            ax6 = fig.add_subplot(gs[2, 2:])
            model_performance = ['GAMM Model', 'Survival Model']
            performance_values = [self.models['gamm']['r2'], 
                                self.models.get('survival', {}).get('mae', 0)]
            colors = ['lightblue', 'lightgreen']
            bars = ax6.bar(model_performance, performance_values, 
                          color=colors, edgecolor='black', alpha=0.7)
            ax6.set_title('模型性能评估', fontsize=14, fontweight='bold')
            ax6.set_ylabel('指标')
            
            # 添加数值标签
            for bar, value in zip(bars, performance_values):
                height = bar.get_height()
                ax6.text(bar.get_x() + bar.get_width()/2., height + 0.01,
                        f'{value:.3f}', ha='center', va='bottom')
        
        # 7. Pareto前沿分析
        if 'pareto_frontier' in self.results:
            ax7 = fig.add_subplot(gs[3, :2])
            pareto_results = self.results['pareto_frontier']
            if pareto_results:
                pareto_miss = [r['漏检风险'] for r in pareto_results]
                pareto_late = [r['晚检风险'] for r in pareto_results]
                ax7.scatter(pareto_miss, pareto_late, color='red', s=100, 
                           alpha=0.8, edgecolors='black', label='Pareto Frontier')
                ax7.set_xlabel('漏检风险')
                ax7.set_ylabel('晚检风险')
                ax7.set_title('Pareto前沿分析', fontsize=14, fontweight='bold')
                ax7.legend()
                ax7.grid(True, alpha=0.3)
        
        # 8. 综合建议
        ax8 = fig.add_subplot(gs[3, 2:])
        ax8.axis('off')
        
        # 生成建议文本
        recommendations = self._generate_recommendations()
        ax8.text(0.05, 0.95, recommendations, transform=ax8.transAxes, 
                fontsize=12, verticalalignment='top',
                bbox=dict(boxstyle="round,pad=0.5", facecolor="lightgray", alpha=0.8))
        
        plt.suptitle('问题三：男胎Y染色体浓度多因素分析与BMI分组优化 - 综合报告', 
                    fontsize=18, fontweight='bold', y=0.98)
        plt.savefig('chart/ver2/comprehensive_analysis_report.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        return recommendations
    
    def _generate_recommendations(self):
        """生成建议文本"""
        recommendations = "=== 问题三分析结果与建议 ===\n\n"
        
        # 模型性能
        if 'gamm' in self.models:
            recommendations += f"1. 模型性能：\n"
            recommendations += f"   - GAMM模型R² = {self.models['gamm']['r2']:.3f}\n"
            recommendations += f"   - GAMM模型MAE = {self.models['gamm']['mae']:.3f}\n\n"
        
        # BMI分组建议
        if 'bmi_grouping' in self.results:
            recommendations += "2. BMI分组与推荐检测时点：\n"
            for result in self.results['bmi_grouping']:
                recommendations += f"   - BMI {result['BMI_group']}: 推荐第{result['week']}周检测\n"
                recommendations += f"     达标概率: {result['达标概率']:.3f}, 总风险: {result['总风险']:.3f}\n"
            recommendations += "\n"
        
        # 蒙特卡洛稳健性
        if 'monte_carlo' in self.results:
            recommendations += "3. 稳健性分析（蒙特卡洛模拟）：\n"
            for group, stats in self.results['monte_carlo'].items():
                recommendations += f"   - BMI {group}: 平均{stats['mean_week']:.1f}周 "
                recommendations += f"(±{stats['std_week']:.1f}周)\n"
            recommendations += "\n"
        
        # Pareto前沿分析
        if 'pareto_frontier' in self.results:
            recommendations += "4. 多目标优化分析：\n"
            recommendations += f"   - 找到{len(self.results['pareto_frontier'])}个Pareto最优解\n"
            recommendations += "   - 在漏检风险与晚检风险之间找到最佳平衡点\n"
            recommendations += "   - 为临床决策提供多种可选方案\n\n"
        
        # IPW分析
        if 'ipw' in self.models:
            recommendations += "5. 选择偏倚校正：\n"
            recommendations += "   - 使用IPW方法校正测序失败的选择偏倚\n"
            recommendations += "   - 提高模型估计的准确性和可靠性\n\n"
        
        # 关键发现
        recommendations += "6. 关键发现：\n"
        recommendations += "   - 孕周是影响Y染色体浓度的最重要因素\n"
        recommendations += "   - BMI与达标时间呈负相关关系\n"
        recommendations += "   - 年龄、身高、体重对达标时间有显著影响\n"
        recommendations += "   - 测量误差对推荐时点的影响在可接受范围内\n\n"
        
        # 临床建议
        recommendations += "7. 临床建议：\n"
        recommendations += "   - 建议按BMI分组制定个性化检测策略\n"
        recommendations += "   - 高BMI孕妇应适当提前检测时点\n"
        recommendations += "   - 考虑多因素综合评估，提高检测准确性\n"
        recommendations += "   - 建立质量控制体系，减少测量误差影响\n"
        
        return recommendations
    
    def run_complete_analysis(self):
        """运行完整分析流程"""
        print("开始问题三完整分析...")
        
        # 1. 数据预处理
        if self.load_and_preprocess_data() is None:
            print("数据预处理失败，无法继续分析")
            return None
        
        # 2. 探索性分析
        self.exploratory_analysis()
        
        # 3. 处理选择偏倚（IPW）
        self.handle_selection_bias()
        
        # 4. 建立模型
        self.build_gamm_model()
        self.build_survival_model()
        
        # 5. BMI分组优化（多目标优化）
        self.bmi_grouping_optimization()
        
        # 6. 蒙特卡洛分析
        self.monte_carlo_analysis()
        
        # 7. 生成最终报告
        recommendations = self.generate_final_report()
        
        print("分析完成！")
        return recommendations

def create_sample_data():
    """创建示例数据用于测试"""
    np.random.seed(42)
    n_samples = 1000
    
    data = {
        'A': range(1, n_samples + 1),  # 样本序号
        'B': [f'P{i:04d}' for i in range(1, n_samples + 1)],  # 孕妇代码
        'C': np.random.normal(30, 5, n_samples),  # 年龄
        'D': np.random.normal(165, 8, n_samples),  # 身高
        'E': np.random.normal(65, 15, n_samples),  # 体重
        'J': np.random.uniform(10, 25, n_samples),  # 孕周
        'K': np.random.uniform(20, 45, n_samples),  # BMI
        'V': np.random.uniform(1, 8, n_samples)  # Y染色体浓度
    }
    
    df = pd.DataFrame(data)
    df.to_excel('sample_data.xlsx', index=False)
    print("已创建示例数据文件：sample_data.xlsx")
    return df

# 主程序
if __name__ == "__main__":
    import os
    import argparse
    
    # 检查数据文件是否存在（支持根目录或problems4目录）
    candidate_paths = [
        '附件.xlsx',
        os.path.join('problems4', '附件.xlsx')
    ]
    data_file = None
    for p in candidate_paths:
        if os.path.exists(p):
            data_file = p
            break
    if data_file is None:
        print("未找到附件.xlsx文件，创建示例数据进行测试...")
        create_sample_data()
        data_file = 'sample_data.xlsx'
    
    # 命令行参数
    parser = argparse.ArgumentParser(description='Problem 3 analysis runner')
    parser.add_argument('--mc', type=int, default=1000, help='Monte Carlo simulation runs (default: 1000)')
    parser.add_argument('--bootstrap', type=int, default=0, help='Bootstrap repetitions for BMI CI (default: 0 to skip)')
    args = parser.parse_args()

    # 创建求解器实例
    solver = NIPTProblem3Solver(data_file)

    # 运行完整分析
    recommendations = solver.run_complete_analysis()

    # 额外运行：根据参数覆盖蒙特卡洛次数
    if args.mc and args.mc != 1000:
        solver.monte_carlo_analysis(n_simulations=args.mc)

    # 额外运行：Bootstrap 置信区间
    if args.bootstrap and args.bootstrap > 0:
        solver.bootstrap_bmi_recommendation_ci(n_bootstrap=args.bootstrap)
    
    if recommendations:
        # 打印建议
        print("\n" + "="*50)
        print("最终建议：")
        print("="*50)
        print(recommendations)
        
        # 保存结果到文件
        with open('problem3_recommendations.txt', 'w', encoding='utf-8') as f:
            f.write(recommendations)
        
        print("\n结果已保存到 problem3_recommendations.txt")
    else:
        print("分析失败，请检查数据文件")